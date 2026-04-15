"""
app/api/batch.py  — Bulk Upload, Template Download & Batch Status
"""
from fastapi import APIRouter, BackgroundTasks, HTTPException, UploadFile, File, Request, Depends
from fastapi.responses import StreamingResponse
from typing import List, Dict, Any
from uuid import uuid4
import json
import io
import logging
from app.core.config import settings
from app.db.session import get_db, SessionLocal
from sqlalchemy.orm import Session

router = APIRouter()
log = logging.getLogger(__name__)

# Optional Redis — degrade gracefully if unavailable
try:
    import redis
    _redis = redis.Redis.from_url(settings.redis_url, decode_responses=True, socket_connect_timeout=1)
    _redis.ping()
    redis_client = _redis
except Exception:
    redis_client = None


# ─── Template columns (51 PINT AE fields) ──────────────────────────────────

TEXT_COLUMNS = {
    "invoice_type_code", "payment_means_type_code", "currency_code",
    "tax_category_code", "transaction_type", "transaction_type_code",
    "seller_country_code", "buyer_country_code", "unit_of_measure", "tax_category",
    "seller_trn", "buyer_trn", "line_id",
    "seller_subdivision", "buyer_subdivision",
    "seller_registration_identifier_type", "buyer_registration_identifier_type",
    "seller_electronic_address", "buyer_electronic_address",
    "seller_bank_iban", "buyer_reference", "invoice_number",
}

TEMPLATE_COLUMNS = [
    # A1 — Invoice header
    "invoice_number", "invoice_date", "payment_due_date", "invoice_type_code",
    "payment_means_type_code", "payment_terms", "transaction_type",
    "transaction_type_code", "currency_code", "tax_category_code", 
    "buyer_reference", "specification_id", "business_process_id",
    "tax_point_date", "order_reference",
    # A2 — Seller
    "seller_name", "seller_trn", "seller_electronic_address",
    "seller_electronic_scheme", "seller_bank_iban", "seller_address",
    "seller_city", "seller_subdivision", "seller_country_code",
    "seller_legal_registration", "seller_registration_identifier_type",
    "seller_postal_code",
    # A3 — Buyer
    "buyer_name", "buyer_trn", "buyer_electronic_address",
    "buyer_electronic_scheme", "buyer_address", "buyer_city",
    "buyer_subdivision", "buyer_country_code",
    "buyer_legal_registration", "buyer_registration_identifier_type",
    "buyer_postal_code",
    # A6 — Line items
    "line_id", "item_name", "item_description", "unit_of_measure",
    "quantity", "unit_price", "line_net_amount",
    "tax_category", "tax_rate", "tax_amount",
    # A4 — Totals
    "total_without_tax", "total_with_tax", "amount_due",
]

SAMPLE_ROWS = [
    # INV-2026-001 — B2B, single line
    {
        "invoice_number": "INV-2026-001", "invoice_date": "2026-04-01",
        "payment_due_date": "2026-04-30", "invoice_type_code": "380",
        "payment_means_type_code": "30", "payment_terms": "Standard 30 Days",
        "transaction_type": "B2B", "transaction_type_code": "10000000",
        "currency_code": "AED", "tax_category_code": "S",
        "buyer_reference": "PO-12345",
        "specification_id": "urn:cen.eu:en16931:2017#compliant#urn:fdc:peppol.eu:2017:poacc:billing:3.0",
        "business_process_id": "urn:fdc:peppol.eu:2017:poacc:billing:01:1.0",
        "seller_name": "Adamas Tech Corp", "seller_trn": "100200300400500",
        "seller_electronic_address": "accounts@adamas-tech.ae",
        "seller_electronic_scheme": "0235", "seller_bank_iban": "AE070331234567890123456",
        "seller_address": "Dubai Silicon Oasis", "seller_city": "Dubai",
        "seller_subdivision": "DU", "seller_country_code": "AE",
        "seller_legal_registration": "L-1002003",
        "seller_registration_identifier_type": "Trade License",
        "buyer_name": "Client Group FZE", "buyer_trn": "100999888777666",
        "buyer_electronic_address": "finance@client-group.ae",
        "buyer_electronic_scheme": "0235",
        "buyer_address": "Abu Dhabi Global Market", "buyer_city": "Abu Dhabi",
        "buyer_subdivision": "AZ", "buyer_country_code": "AE",
        "buyer_legal_registration": "L-9988776",
        "buyer_registration_identifier_type": "Trade License",
        "line_id": "1", "item_name": "Consulting Services",
        "item_description": "IT Strategy & Architecture Consulting",
        "unit_of_measure": "EA", "quantity": 10, "unit_price": 500.00,
        "line_net_amount": 5000.00, "tax_category": "S", "tax_rate": 0.05,
        "tax_amount": 250.00, "total_without_tax": 5000.00,
        "total_with_tax": 5250.00, "amount_due": 5250.00,
    },
    # INV-2026-002 — B2B, line 2 (same invoice, multi-line)
    {
        "invoice_number": "INV-2026-001", "invoice_date": "2026-04-01",
        "payment_due_date": "2026-04-30", "invoice_type_code": "380",
        "payment_means_type_code": "30", "payment_terms": "Standard 30 Days",
        "transaction_type": "B2B", "transaction_type_code": "10000000",
        "currency_code": "AED", "tax_category_code": "S",
        "buyer_reference": "PO-12345",
        "specification_id": "urn:cen.eu:en16931:2017#compliant#urn:fdc:peppol.eu:2017:poacc:billing:3.0",
        "business_process_id": "urn:fdc:peppol.eu:2017:poacc:billing:01:1.0",
        "seller_name": "Adamas Tech Corp", "seller_trn": "100200300400500",
        "seller_electronic_address": "accounts@adamas-tech.ae",
        "seller_electronic_scheme": "0235", "seller_bank_iban": "AE070331234567890123456",
        "seller_address": "Dubai Silicon Oasis", "seller_city": "Dubai",
        "seller_subdivision": "DU", "seller_country_code": "AE",
        "seller_legal_registration": "L-1002003",
        "seller_registration_identifier_type": "Trade License",
        "buyer_name": "Client Group FZE", "buyer_trn": "100999888777666",
        "buyer_electronic_address": "finance@client-group.ae",
        "buyer_electronic_scheme": "0235",
        "buyer_address": "Abu Dhabi Global Market", "buyer_city": "Abu Dhabi",
        "buyer_subdivision": "AZ", "buyer_country_code": "AE",
        "buyer_legal_registration": "L-9988776",
        "buyer_registration_identifier_type": "Trade License",
        "line_id": "2", "item_name": "Cloud Infrastructure Setup",
        "item_description": "AWS/Azure architecture design and deployment",
        "unit_of_measure": "EA", "quantity": 1, "unit_price": 2500.00,
        "line_net_amount": 2500.00, "tax_category": "S", "tax_rate": 0.05,
        "tax_amount": 125.00, "total_without_tax": 7500.00,
        "total_with_tax": 7875.00, "amount_due": 7875.00,
    },
    # INV-2026-003 — B2C
    {
        "invoice_number": "INV-2026-003", "invoice_date": "2026-04-02",
        "payment_due_date": "2026-04-02", "invoice_type_code": "380",
        "payment_means_type_code": "10", "payment_terms": "Cash on Delivery",
        "transaction_type": "B2C", "transaction_type_code": "01000000",
        "currency_code": "AED", "tax_category_code": "S",
        "buyer_reference": "POS-CASH-001",
        "specification_id": "urn:cen.eu:en16931:2017#compliant#urn:fdc:peppol.eu:2017:poacc:billing:3.0",
        "business_process_id": "urn:fdc:peppol.eu:2017:poacc:billing:01:1.0",
        "seller_name": "Adamas Tech Corp", "seller_trn": "100200300400500",
        "seller_electronic_address": "pos@adamas-tech.ae",
        "seller_electronic_scheme": "0235", "seller_bank_iban": "AE070331234567890123456",
        "seller_address": "Dubai Silicon Oasis", "seller_city": "Dubai",
        "seller_subdivision": "DU", "seller_country_code": "AE",
        "seller_legal_registration": "L-1002003",
        "seller_registration_identifier_type": "Trade License",
        "buyer_name": "Individual Customer", "buyer_trn": "",
        "buyer_electronic_address": "consumer@example.com",
        "buyer_electronic_scheme": "0235",
        "buyer_address": "Sharjah City", "buyer_city": "Sharjah",
        "buyer_subdivision": "SH", "buyer_country_code": "AE",
        "buyer_legal_registration": "", "buyer_registration_identifier_type": "",
        "line_id": "1", "item_name": "Software License",
        "item_description": "Annual Software Subscription",
        "unit_of_measure": "EA", "quantity": 2, "unit_price": 100.00,
        "line_net_amount": 200.00, "tax_category": "S", "tax_rate": 0.05,
        "tax_amount": 10.00, "total_without_tax": 200.00,
        "total_with_tax": 210.00, "amount_due": 210.00,
    },
]

# ─── Download Template (XLSX) ───────────────────────────────────────────────

@router.get("/download-template")
async def download_template():
    """
    Returns a professional Excel/XLSX template for bulk uploads.
    All 51 mandatory PINT AE fields, 3 sample invoices (including a multi-line example).
    Text columns are pre-formatted as '@' to prevent TRN integer-casting.
    """
    from openpyxl import Workbook
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "PINT_AE_Invoices"

    # Instructions sheet
    ws_info = wb.create_sheet("Instructions")
    ws_info["A1"] = "UAE PINT AE Bulk Upload Template"
    ws_info["A2"] = "Version: 2.0 | Format: One row per invoice line item"
    ws_info["A3"] = "Multi-line invoices: repeat the invoice_number with same header data, change line_id"
    ws_info["A5"] = "Key Rules:"
    rules = [
        "seller_trn / buyer_trn: Must be exactly 15 digits (numeric only)",
        "invoice_date / payment_due_date: YYYY-MM-DD format",
        "tax_rate: Use decimal (0.05 = 5%). Standard-rated must be 0.05",
        "transaction_type: B2B | B2C | B2G (B2B requires buyer_trn)",
        "invoice_type_code: 380=Invoice, 381=Credit Note, 383=Debit Note",
        "currency_code: AED (primary), USD, EUR (supported)",
        "unit_of_measure: UN/ECE Rec 20 codes (EA=Each, HUR=Hour, DAY=Day)",
        "tax_category: S=Standard(5%), Z=Zero-rated, E=Exempt, O=Outside scope",
    ]
    for i, rule in enumerate(rules, start=6):
        ws_info[f"A{i}"] = f"• {rule}"

    # Style definitions
    header_fill = PatternFill("solid", fgColor="1A2340")
    header_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    sample_fill_b2b = PatternFill("solid", fgColor="EAF4FF")
    sample_fill_b2c = PatternFill("solid", fgColor="F0FFF0")
    thin_border = Border(
        bottom=Side(style="thin", color="C0CCE0"),
        right=Side(style="thin", color="E3EAF7"),
    )
    section_fills = {
        # A1 Invoice header - light blue
        range(0, 13): PatternFill("solid", fgColor="DBEAFE"),
        # A2 Seller - light green
        range(13, 24): PatternFill("solid", fgColor="DCFCE7"),
        # A3 Buyer - light yellow
        range(24, 34): PatternFill("solid", fgColor="FEF9C3"),
        # A6 Lines - light purple
        range(34, 44): PatternFill("solid", fgColor="EDE9FE"),
        # A4 Totals - light orange
        range(44, 47): PatternFill("solid", fgColor="FFEDD5"),
    }

    # Header row with color by section
    for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = max(len(col_name) + 3, 14)

        # Pre-format columns as Text to prevent Excel mangling
        if col_name in TEXT_COLUMNS:
            for row_idx in range(2, 502):
                ws.cell(row=row_idx, column=col_idx).number_format = "@"

    ws.row_dimensions[1].height = 32
    ws.freeze_panes = "A2"

    # Sample rows
    for row_idx, row_data in enumerate(SAMPLE_ROWS, start=2):
        fill = sample_fill_b2c if row_data.get("transaction_type") == "B2C" else sample_fill_b2b
        for col_idx, col_name in enumerate(TEMPLATE_COLUMNS, start=1):
            val = row_data.get(col_name, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.fill = fill
            cell.border = thin_border
            if col_name in TEXT_COLUMNS:
                cell.number_format = "@"
                cell.value = str(val) if val != "" else ""

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=PINT_AE_Bulk_Template.xlsx"},
    )


# ─── Download CSV Template ──────────────────────────────────────────────────

@router.get("/download-template-csv")
async def download_template_csv():
    """Returns the same template as CSV for direct upload testing."""
    import csv as csvlib

    output = io.StringIO()
    writer = csvlib.DictWriter(output, fieldnames=TEMPLATE_COLUMNS, extrasaction="ignore")
    writer.writeheader()
    for row in SAMPLE_ROWS:
        writer.writerow({k: row.get(k, "") for k in TEMPLATE_COLUMNS})

    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=PINT_AE_Bulk_Template.csv"},
    )


# ─── Bulk Upload ────────────────────────────────────────────────────────────

@router.post("/ingest-bulk")
@router.post("/upload-bulk")
@router.post("/upload-excel")
async def upload_bulk(
    request: Request,
    background_tasks: BackgroundTasks,
    file: UploadFile = File(...),
    full_pipeline: bool = False,
):
    filename = file.filename.lower()
    if not any(filename.endswith(ext) for ext in [".xlsx", ".xls", ".csv"]):
        raise HTTPException(400, "Supported formats: .xlsx, .xls, .csv")

    contents = await file.read()
    tenant_id = getattr(request.state, "tenant_id", "anonymous")
    batch_id = f"BATCH-{uuid4().hex[:8].upper()}"

    from app.db.models import ETLJob, ETLJobStatus
    from app.etl.tasks.extract import extract_excel

    db = SessionLocal()
    try:
        source_format = "excel" if filename.endswith((".xlsx", ".xls")) else "csv"
        job = ETLJob(
            batch_id=batch_id,
            tenant_id=tenant_id,
            source_filename=file.filename,
            source_format=source_format,
            status=ETLJobStatus.QUEUED.value,
        )
        db.add(job)
        db.commit()
        db.refresh(job)

        file_hex = contents.hex()
        extract_excel.delay(job.id, file_hex, file.filename, tenant_id=tenant_id, full_pipeline=full_pipeline)

        return {
            "batch_id": batch_id,
            "job_id": job.id,
            "status": "QUEUED",
            "message": f"Processing {file.filename}. Poll /api/v1/batch-status/{batch_id} for results.",
            "poll_url": f"/api/v1/batch-status/{batch_id}",
        }
    except Exception as e:
        db.rollback()
        log.error(f"Failed to initiate bulk upload: {e}")
        raise HTTPException(500, f"Upload failed: {str(e)}")
    finally:
        db.close()


# ─── Batch API Submit (JSON payloads) ──────────────────────────────────────

@router.post("/batch-validate")
async def batch_validate(
    request: Request,
    payloads: List[Dict[Any, Any]],
    full_pipeline: bool = False,
):
    if not payloads:
        raise HTTPException(400, "No payloads provided")
    if len(payloads) > 500:
        raise HTTPException(400, "Maximum 500 invoices per batch")

    tenant_id = getattr(request.state, "tenant_id", "anonymous")
    batch_id = f"BATCH-{uuid4().hex[:8].upper()}"

    from app.db.models import ETLJob, ETLJobStatus
    from app.etl.tasks.transform import transform_batch

    db = SessionLocal()
    try:
        job = ETLJob(
            batch_id=batch_id,
            tenant_id=tenant_id,
            job_type="api_batch",
            status=ETLJobStatus.QUEUED.value,
            total_rows=len(payloads),
        )
        db.add(job)
        db.commit()
        db.refresh(job)

        transform_batch.delay(job.id, payloads, tenant_id=tenant_id, full_pipeline=full_pipeline)

        return {
            "batch_id": batch_id,
            "job_id": job.id,
            "status": "ACCEPTED",
            "total": len(payloads),
            "poll_url": f"/api/v1/batch-status/{batch_id}",
        }
    except Exception as e:
        db.rollback()
        raise HTTPException(500, f"Batch validation failed: {str(e)}")
    finally:
        db.close()


# ─── Batch / Job Status ─────────────────────────────────────────────────────

@router.get("/batch-status/{batch_id}")
async def batch_status(request: Request, batch_id: str, db: Session = Depends(get_db)):
    if redis_client:
        cached = redis_client.get(f"batch:{batch_id}")
        if cached:
            try:
                data = json.loads(cached)
                # If COMPLETE, we MUST bypass cache to get the full results list from DB
                if data.get("status") not in ["COMPLETE", "FAILED", "PARTIAL"]:
                    return data
            except:
                pass

    # 2. Database lookup
    from app.db.models import ETLJob, ValidationRun, ETLRowError
    job = db.query(ETLJob).filter(ETLJob.batch_id == batch_id).first()
    if not job:
        raise HTTPException(404, f"Batch '{batch_id}' not found")

    results = []
    runs = (
        db.query(ValidationRun)
        .filter(ValidationRun.etl_job_id == job.id)
        .order_by(ValidationRun.invoice_number)
        .all()
    )
    for r in runs:
        results.append({
            "invoice_number": r.invoice_number,
            "is_valid": r.is_valid,
            "errors": r.errors_json or [],
            "pass_percentage": r.pass_percentage,
            "total_errors": r.total_errors,
            "asp_status": r.asp_status,
            "asp_clearance_id": r.asp_clearance_id,
            "phase": r.phase,
            "error_phase": r.error_phase,
            "row_number": getattr(r, "row_number", None),
        })

    # Add transform/extract phase errors
    row_errors = db.query(ETLRowError).filter(ETLRowError.etl_job_id == job.id).all()
    for re in row_errors:
        results.append({
            "invoice_number": re.invoice_number or f"Row {re.row_number}",
            "is_valid": False,
            "errors": [{"field": re.error_type, "error": re.error_message}],
            "pass_percentage": 0,
            "total_errors": 1,
            "phase": "extract_transform",
            "row_number": re.row_number,
        })

    total = job.total_rows or 0
    done = (job.processed_rows or 0) + (job.failed_rows or 0)

    return {
        "status": job.status,
        "total": total,
        "done": done,
        "valid": job.valid_rows or 0,
        "invalid": job.invalid_rows or 0,
        "batch_id": job.batch_id,
        "job_id": job.id,
        "source_filename": job.source_filename,
        "results": results,
        "error_message": job.error_message,
    }


# ─── ETL Job Detail / List ──────────────────────────────────────────────────

@router.get("/etl-jobs/{job_id}")
async def get_etl_job(job_id: str, request: Request, db: Session = Depends(get_db)):
    from app.db.models import ETLJob
    job = db.query(ETLJob).filter(ETLJob.id == job_id).first()
    if not job:
        raise HTTPException(404, "ETL job not found")
    return {
        "id": job.id,
        "batch_id": job.batch_id,
        "status": job.status,
        "source_filename": job.source_filename,
        "total_rows": job.total_rows,
        "processed_rows": job.processed_rows,
        "valid_rows": job.valid_rows,
        "invalid_rows": job.invalid_rows,
        "failed_rows": job.failed_rows,
        "progress_pct": round(job.processed_rows / job.total_rows * 100, 1) if job.total_rows else 0,
        "duration_ms": job.duration_ms,
        "started_at": job.started_at,
        "completed_at": job.completed_at,
        "created_at": job.created_at,
        "error_message": job.error_message,
    }


@router.get("/etl-jobs")
async def list_etl_jobs(
    request: Request,
    db: Session = Depends(get_db),
    skip: int = 0,
    limit: int = 20,
):
    from app.db.models import ETLJob
    from sqlalchemy import desc
    tenant_id = getattr(request.state, "tenant_id", "anonymous")
    jobs = (
        db.query(ETLJob)
        .filter(ETLJob.tenant_id == tenant_id)
        .order_by(desc(ETLJob.created_at))
        .offset(skip)
        .limit(limit)
        .all()
    )
    return [
        {
            "id": j.id,
            "batch_id": j.batch_id,
            "status": j.status,
            "source_filename": j.source_filename,
            "total_rows": j.total_rows,
            "valid_rows": j.valid_rows,
            "invalid_rows": j.invalid_rows,
            "progress_pct": round(j.processed_rows / j.total_rows * 100, 1) if j.total_rows else 0,
            "created_at": j.created_at,
        }
        for j in jobs
    ]
